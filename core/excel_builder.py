"""
Gera o arquivo Excel de cálculo de importação a partir dos dados estruturados.
Cria 5 abas:
  1. Inputs           — dados do espelho preenchidos
  2. Cálculo          — memória de cálculo automática
  3. Fórmulas Protheus — fórmulas para o Configurador de Tributos
  4. Conferência      — comparativo com os totais do espelho
  5. Guia             — instruções e referência

Toda a estrutura segue o Template_Calculo_Importacao_Protheus já validado.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================
# Estilos padrão
# ============================================================
ARIAL = "Arial"
CONSOLAS = "Consolas"

H1     = Font(name=ARIAL, size=14, bold=True, color="FFFFFF")
H2     = Font(name=ARIAL, size=11, bold=True, color="1F3864")
HDR    = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
BLK    = Font(name=ARIAL, size=10)
BLU    = Font(name=ARIAL, size=10, color="0000FF")
BOLD   = Font(name=ARIAL, size=10, bold=True)
SEC    = Font(name=ARIAL, size=10, bold=True, color="1F3864")
GRN    = Font(name=ARIAL, size=10, bold=True, color="008000")
MONO   = Font(name=CONSOLAS, size=10, color="1F3864")

NAVY   = PatternFill("solid", start_color="1F3864")
MID    = PatternFill("solid", start_color="2F5496")
SECF   = PatternFill("solid", start_color="E7EDF7")
GRNF   = PatternFill("solid", start_color="E2EFDA")
INPF   = PatternFill("solid", start_color="FFF7E5")
CODEF  = PatternFill("solid", start_color="F2F8FD")

thin   = Side(style="thin", color="B4C6E7")
BORD   = Border(left=thin, right=thin, top=thin, bottom=thin)

NUM    = '#,##0.00;(#,##0.00);"-"'
PCT2   = "0.00%"
PCT4   = "0.0000%"
CENT   = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


N_ITENS = 10
ITEM_COLS = [get_column_letter(i) for i in range(3, 3 + N_ITENS)]   # C..L
TOT_COL   = get_column_letter(3 + N_ITENS)                          # M
OBS_COL   = get_column_letter(4 + N_ITENS)                          # N


# ============================================================
# Aba 1: Inputs
# ============================================================
def _build_inputs(wb: Workbook, dados: dict) -> dict:
    """Cria a aba 1 e retorna as referências de linha usadas."""
    ws = wb.active
    ws.title = "1. Inputs"
    ws.sheet_view.showGridLines = False

    ws.merge_cells(f"B2:{OBS_COL}2")
    c = ws["B2"]
    c.value = "INPUTS POR ITEM — DADOS DO ESPELHO DA NF DE IMPORTAÇÃO"
    c.font, c.fill, c.alignment = H1, NAVY, CENT
    ws.row_dimensions[2].height = 24

    ws.merge_cells(f"B3:{OBS_COL}3")
    ws["B3"].value = (
        "Células laranja = inputs (já preenchidos a partir do espelho). "
        "Ajuste o que precisar; a aba 2 recalcula automaticamente."
    )
    ws["B3"].font = Font(name=ARIAL, size=9, italic=True, color="595959")

    # Cabeçalho do processo
    cab = dados.get("cabecalho", {})
    ws["B5"].value, ws["B5"].font = "IDENTIFICAÇÃO DO PROCESSO", H2
    ws.merge_cells("B5:F5")

    def _kv(row, label_col, val_col, label, value, fmt=None):
        ws.cell(row=row, column=label_col, value=label).font = BLK
        ws.cell(row=row, column=label_col).border = BORD
        c = ws.cell(row=row, column=val_col, value=value)
        c.fill, c.font, c.border = INPF, BLU, BORD
        if fmt: c.number_format = fmt
        return c

    _kv(6, 2, 3, "Nº do Processo", cab.get("processo", ""))
    _kv(6, 4, 5, "DUIMP", cab.get("duimp", ""))
    ws.merge_cells("E6:F6")

    _kv(7, 2, 3, "Fornecedor", cab.get("fornecedor", ""))
    ws.merge_cells("C7:F7")

    c = _kv(8, 2, 3, "Modal", cab.get("modal", "AÉREO"))
    dv = DataValidation(type="list", formula1='"AÉREO,MARÍTIMO,RODOVIÁRIO,FERROVIÁRIO"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(c)
    _kv(8, 4, 5, "Taxa câmbio USD", cab.get("taxa_cambio", 0.0), fmt="0.0000000")
    ws.merge_cells("E8:F8")

    # Cabeçalho da grade
    ws.cell(row=10, column=2, value="ITEM").font = H2
    for j in range(N_ITENS):
        c = ws.cell(row=10, column=3 + j, value=f"Item {j+1}")
        c.font, c.fill, c.alignment, c.border = HDR, MID, CENT, BORD
    for j, t in enumerate(["Total", "Observações"]):
        c = ws.cell(row=10, column=3 + N_ITENS + j, value=t)
        c.font, c.fill, c.alignment, c.border = HDR, MID, CENT, BORD
    ws.row_dimensions[10].height = 22

    def _secao(r, t):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4 + N_ITENS)
        c = ws.cell(row=r, column=2, value=t); c.font, c.fill = SEC, SECF

    def _input(r, label, valores, fmt, obs="", soma=True):
        ws.cell(row=r, column=2, value=label).font = BLK
        ws.cell(row=r, column=2).border = BORD
        for j in range(N_ITENS):
            c = ws.cell(row=r, column=3 + j)
            if j < len(valores) and valores[j] is not None:
                c.value = valores[j]
            c.fill, c.font, c.number_format, c.border = INPF, BLU, fmt, BORD
        if soma:
            c = ws.cell(row=r, column=3 + N_ITENS,
                        value=f"=SUM({ITEM_COLS[0]}{r}:{ITEM_COLS[-1]}{r})")
            c.font, c.number_format, c.border = BOLD, fmt, BORD
        else:
            ws.cell(row=r, column=3 + N_ITENS, value="").border = BORD
        co = ws.cell(row=r, column=4 + N_ITENS, value=obs)
        co.font = Font(name=ARIAL, size=9, italic=True, color="595959")
        co.border = BORD

    itens = dados.get("itens", [])
    def _vals(key, default=None):
        out = []
        for i in range(N_ITENS):
            if i < len(itens):
                out.append(itens[i].get(key, default))
            else:
                out.append(None)
        return out

    r = 11
    _secao(r, "— IDENTIFICAÇÃO E QUANTIDADES —"); r += 1
    refs = {}
    refs["desc"] = r; _input(r, "Descrição do produto", _vals("descricao"), "@", "livre", soma=False); r += 1
    refs["ncm"]  = r; _input(r, "NCM", _vals("ncm"), "@", "ex.: 8443.32.99", soma=False); r += 1
    refs["qt"]   = r; _input(r, "Quantidade", _vals("qtd"), "#,##0", "soma das qtds."); r += 1
    refs["vu"]   = r; _input(r, "Valor unitário (R$)", _vals("vunit"), '#,##0.0000', "preço unitário do espelho", soma=False); r += 1

    _secao(r, "— VALORES POR ITEM (R$) —"); r += 1
    refs["cif"]  = r; _input(r, "CIF / VMLD do item (R$)", _vals("cif"), NUM, "inclui frete e seguro int. já rateados"); r += 1
    # Despesas que ENTRAM na base do ICMS: Siscomex + AFRMM (rateado) por item
    refs["outd"] = r; _input(r, "Outras despesas integrantes da base ICMS (R$)", _vals("outras_base_icms"), NUM, "Siscomex + AFRMM rateados"); r += 1
    # Despesas que NÃO entram na base: Desp Ac líquido de AFRMM (que já está embutido nele)
    refs["outv"] = r; _input(r, "Outras despesas SEM influência no ICMS (R$)", _vals("outras_sem_icms"), NUM, "Desp Ac do espelho líquido de AFRMM"); r += 1

    _secao(r, "— ALÍQUOTAS POR ITEM —"); r += 1
    refs["aii"]  = r; _input(r, "Alíquota II (%)", _vals("aliq_ii"), PCT2, "0% se Ex-tarifário", soma=False); r += 1
    refs["aipi"] = r; _input(r, "Alíquota IPI (%)", _vals("aliq_ipi"), PCT2, "conforme NCM (TIPI)", soma=False); r += 1
    refs["apis"] = r; _input(r, "Alíquota PIS-importação (%)", _vals("aliq_pis"), PCT2, "padrão 2,10%", soma=False); r += 1
    refs["acof"] = r; _input(r, "Alíquota COFINS-importação (%)", _vals("aliq_cofins"), PCT2, "9,65% ou 10,25% (majorada)", soma=False); r += 1
    # Alíquota cheia interna — vem 0,18 por default (MG); pode mudar manualmente
    cheias = [0.18 if i < len(itens) else None for i in range(N_ITENS)]
    refs["ach"] = r; _input(r, "Alíquota ICMS interna cheia (%)", cheias, PCT2, "ex.: MG 18%", soma=False); r += 1
    refs["car"] = r; _input(r, "Carga efetiva ICMS (%)", _vals("aliq_icms"), PCT2, "se = cheia → sem redução; < cheia → Conv. 52/91 etc.", soma=False); r += 1

    # Larguras
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 46
    for col in ITEM_COLS: ws.column_dimensions[col].width = 12.5
    ws.column_dimensions[TOT_COL].width = 13
    ws.column_dimensions[OBS_COL].width = 36
    ws.freeze_panes = "C11"

    return refs


# ============================================================
# Aba 2: Cálculo
# ============================================================
def _build_calculo(wb: Workbook, refs_inputs: dict) -> dict:
    ws = wb.create_sheet("2. Cálculo")
    ws.sheet_view.showGridLines = False

    ws.merge_cells(f"B2:{OBS_COL}2")
    c = ws["B2"]; c.value = "CÁLCULO DOS IMPOSTOS DE IMPORTAÇÃO — MEMÓRIA AUTOMÁTICA"
    c.font, c.fill, c.alignment = H1, NAVY, CENT
    ws.row_dimensions[2].height = 24

    ws.merge_cells(f"B3:{OBS_COL}3")
    ws["B3"].value = "Recalcula automaticamente a partir da aba '1. Inputs'."
    ws["B3"].font = Font(name=ARIAL, size=9, italic=True, color="595959")

    # Header
    for j, h in enumerate(["Etapa do cálculo"] + [f"Item {i+1}" for i in range(N_ITENS)] + ["TOTAL", "Observação"], start=2):
        c = ws.cell(row=5, column=j, value=h); c.font, c.fill, c.alignment, c.border = HDR, MID, CENT, BORD
    ws.row_dimensions[5].height = 22

    def _sec(r, t):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4 + N_ITENS)
        ws.cell(row=r, column=2, value=t).font = SEC
        ws.cell(row=r, column=2).fill = SECF

    def _linha(r, lab, fmt_template, fmt, obs="", destaque=False, soma=True):
        ws.cell(row=r, column=2, value=lab).font = BLK
        ws.cell(row=r, column=2).border = BORD
        for j, col in enumerate(ITEM_COLS):
            c = ws.cell(row=r, column=3 + j, value=fmt_template.format(col=col))
            c.font = BOLD if destaque else BLK
            c.number_format, c.border = fmt, BORD
        if soma:
            c = ws.cell(row=r, column=3 + N_ITENS, value=f"=SUM({ITEM_COLS[0]}{r}:{ITEM_COLS[-1]}{r})")
            c.font, c.number_format, c.border = BOLD, fmt, BORD
        else:
            ws.cell(row=r, column=3 + N_ITENS, value="").border = BORD
        co = ws.cell(row=r, column=4 + N_ITENS, value=obs)
        co.font = Font(name=ARIAL, size=9, italic=True, color="595959")
        co.border = BORD
        if destaque:
            for col in range(2, 5 + N_ITENS): ws.cell(row=r, column=col).fill = GRNF

    def _replica(r, lab, ref_row, fmt, obs=""):
        ws.cell(row=r, column=2, value=lab).font = BLK
        ws.cell(row=r, column=2).border = BORD
        for j, col in enumerate(ITEM_COLS):
            c = ws.cell(row=r, column=3 + j, value=f"='1. Inputs'!{col}{ref_row}")
            c.font, c.number_format, c.border = BLK, fmt, BORD
        ws.cell(row=r, column=3 + N_ITENS).border = BORD
        co = ws.cell(row=r, column=4 + N_ITENS, value=obs)
        co.font = Font(name=ARIAL, size=9, italic=True, color="595959"); co.border = BORD

    R = refs_inputs
    r = 6
    _sec(r, "1) CIF / VMLD por item"); r += 1
    _linha(r, "CIF / VMLD do item (R$)", f"='1. Inputs'!{{col}}{R['cif']}", NUM, "replicado dos inputs"); r_cif = r; r += 1

    _sec(r, "2) II — Imposto de Importação  (base = CIF)"); r += 1
    _replica(r, "Alíquota II", R["aii"], PCT2); r_aii = r; r += 1
    _linha(r, "II (R$) = CIF × alíq.", f"=ROUND({{col}}{r_cif}*{{col}}{r_aii},2)", NUM, "0,00 se Ex-tarifário"); r_ii = r; r += 1

    _sec(r, "3) IPI  (base = CIF + II)"); r += 1
    _linha(r, "Base IPI (R$) = CIF + II", f"={{col}}{r_cif}+{{col}}{r_ii}", NUM); r_bipi = r; r += 1
    _replica(r, "Alíquota IPI", R["aipi"], PCT2); r_aipi = r; r += 1
    _linha(r, "IPI (R$) = Base × alíq.", f"=ROUND({{col}}{r_bipi}*{{col}}{r_aipi},2)", NUM); r_ipi = r; r += 1

    _sec(r, "4) PIS/COFINS-Importação  (base = CIF)"); r += 1
    _replica(r, "Alíquota PIS-importação", R["apis"], PCT2); r_apis = r; r += 1
    _linha(r, "PIS (R$) = CIF × alíq.", f"=ROUND({{col}}{r_cif}*{{col}}{r_apis},2)", NUM); r_pis = r; r += 1
    _replica(r, "Alíquota COFINS-importação", R["acof"], PCT2); r_acof = r; r += 1
    _linha(r, "COFINS (R$) = CIF × alíq.", f"=ROUND({{col}}{r_cif}*{{col}}{r_acof},2)", NUM); r_cof = r; r += 1

    _sec(r, "5) DESPESAS NA BASE DO ICMS (Siscomex + AFRMM + outras integrantes)"); r += 1
    _linha(r, "Outras despesas integrantes da base ICMS (R$)", f"='1. Inputs'!{{col}}{R['outd']}", NUM, "soma rateada por item"); r_outd = r; r += 1

    _sec(r, "6) ICMS-Importação  (gross-up 'por dentro')"); r += 1
    _linha(r, "Numerador = CIF + II + IPI + PIS + COFINS + Outras desp.",
           f"={{col}}{r_cif}+{{col}}{r_ii}+{{col}}{r_ipi}+{{col}}{r_pis}+{{col}}{r_cof}+{{col}}{r_outd}", NUM); r_num = r; r += 1
    _replica(r, "Alíquota interna cheia (UF)", R["ach"], PCT2); r_ach = r; r += 1
    _linha(r, "Divisor do ICMS (1 − alíq. cheia)",
           f"=1-{{col}}{r_ach}", PCT2, "ex.: MG 18% → 82%", soma=False); r_div = r; r += 1
    _replica(r, "Carga efetiva do ICMS", R["car"], PCT2, "se < cheia, há redução de base"); r_car = r; r += 1
    _linha(r, "Fator de redução de base equivalente (1 − carga/cheia)",
           f"=IFERROR(1-{{col}}{r_car}/{{col}}{r_ach},0)", PCT4, "0% se sem redução", soma=False); r += 1
    _linha(r, "Fator de base reduzida restante (carga/cheia)",
           f"=IFERROR({{col}}{r_car}/{{col}}{r_ach},0)", PCT4, "100% se sem redução", soma=False); r += 1
    _linha(r, "BASE DO ICMS = Numerador ÷ (1 − alíq. cheia)",
           f"=IFERROR(ROUND({{col}}{r_num}/{{col}}{r_div},2),0)", NUM, "base 'cheia' antes da redução"); r_bas = r; r += 1
    _linha(r, "ICMS DESTACADO (R$) = Base × carga efetiva",
           f"=ROUND({{col}}{r_bas}*{{col}}{r_car},2)", NUM, "destacado na NF e a recolher", destaque=True); r_icm = r; r += 1
    _linha(r, "ICMS 'valor devido' Portal = Base × alíq. cheia",
           f"=ROUND({{col}}{r_bas}*{{col}}{r_ach},2)", NUM, "informativo no extrato da DUIMP"); r_dev = r; r += 1

    _sec(r, "7) FECHAMENTO DA NF"); r += 1
    _linha(r, "vProd do item (= CIF + II)", f"={{col}}{r_cif}+{{col}}{r_ii}", NUM, "soma = Vl Total Produtos"); r_vp = r; r += 1
    _linha(r, "Outras despesas SEM influência no ICMS (R$)", f"='1. Inputs'!{{col}}{R['outv']}", NUM, "compõe vNF"); r_outv = r; r += 1
    _linha(r, "vNF do item = vProd + Outras (base ICMS) + Outras (s/ ICMS) + IPI + ICMS",
           f"={{col}}{r_vp}+{{col}}{r_outd}+{{col}}{r_outv}+{{col}}{r_ipi}+{{col}}{r_icm}", NUM, "soma = Vl Total Nota", destaque=True); r_vnf = r

    # Larguras
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 52
    for col in ITEM_COLS: ws.column_dimensions[col].width = 12.5
    ws.column_dimensions[TOT_COL].width = 13.5
    ws.column_dimensions[OBS_COL].width = 42
    ws.freeze_panes = "C6"

    return dict(cif=r_cif, ii=r_ii, ipi=r_ipi, pis=r_pis, cof=r_cof, outd=r_outd,
                num=r_num, ach=r_ach, car=r_car, bas=r_bas, icm=r_icm, dev=r_dev,
                vp=r_vp, outv=r_outv, vnf=r_vnf, aii=r_aii, aipi=r_aipi, apis=r_apis, acof=r_acof, bipi=r_bipi)


# ============================================================
# Aba 3: Fórmulas Protheus
# ============================================================
def _build_formulas(wb: Workbook, refs_inputs: dict) -> None:
    ws = wb.create_sheet("3. Fórmulas Protheus")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:H2")
    c = ws["B2"]; c.value = "FÓRMULAS GERADAS PARA O CONFIGURADOR DE TRIBUTOS (FISA170)"
    c.font, c.fill, c.alignment = H1, NAVY, CENT
    ws.row_dimensions[2].height = 24

    ws.merge_cells("B3:H3")
    ws["B3"].value = ("Copie-cole no Editor de Fórmulas do Configurador. "
                     "Ajuste os códigos das regras (EII001, EIPI01, AIC001) se na sua codificação forem outros.")
    ws["B3"].font = Font(name=ARIAL, size=9, italic=True, color="595959")

    # Códigos das regras
    ws["B5"].value, ws["B5"].font = "CÓDIGOS DAS REGRAS NO SEU AMBIENTE", H2
    ws.merge_cells("B5:H5")
    cfgs = [
        ("Regra de Valor do II",          "EII001", "Código da Regra de Cálculo do II em entradas"),
        ("Regra de Valor do IPI",         "EIPI01", "Código da Regra de Cálculo do IPI em entradas"),
        ("Regra de Valor do PIS-imp",     "EPIS01", "Código da Regra do PIS-importação"),
        ("Regra de Valor do COFINS-imp",  "ECOF01", "Código da Regra do COFINS-importação"),
        ("Regra de Alíquota do ICMS",     "AIC001", "Código da Regra de Alíquota do ICMS interna"),
    ]
    for i, (lab, val, obs) in enumerate(cfgs):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).font = BLK; ws.cell(row=r, column=2).border = BORD
        c = ws.cell(row=r, column=3, value=val); c.font, c.fill, c.alignment, c.border = BLU, INPF, CENT, BORD
        c = ws.cell(row=r, column=4, value=obs); c.font = Font(name=ARIAL, size=9, italic=True, color="595959"); c.border = BORD
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
    ref_cod = dict(ii="C6", ipi="C7", pis="C8", cof="C9", icms="C10")

    # Análise por item
    r = 12
    ws.merge_cells(f"B{r}:H{r}")
    ws.cell(row=r, column=2, value="DETECÇÃO AUTOMÁTICA DO MÉTODO POR ITEM").font = H2; r += 1
    ws.merge_cells(f"B{r}:H{r}")
    ws.cell(row=r, column=2, value="A planilha identifica se há redução de base e qual fórmula aplicar.").font = Font(name=ARIAL, size=9, italic=True, color="595959"); r += 1

    hdr = ["Item", "NCM", "Alíq. cheia", "Carga efetiva", "Tem redução?",
           "Fator restante (carga/cheia)", "Método aplicável"]
    for j, h in enumerate(hdr, start=2):
        c = ws.cell(row=r, column=j, value=h); c.font, c.fill, c.alignment, c.border = HDR, MID, CENT, BORD
    ws.row_dimensions[r].height = 30
    r += 1

    for i, col in enumerate(ITEM_COLS):
        rr = r + i
        ws.cell(row=rr, column=2, value=f"Item {i+1}").font = BLK
        ws.cell(row=rr, column=2).border = BORD; ws.cell(row=rr, column=2).alignment = CENT
        ws.cell(row=rr, column=3, value=f"='1. Inputs'!{col}{refs_inputs['ncm']}").font = BLK
        ws.cell(row=rr, column=3).border = BORD; ws.cell(row=rr, column=3).alignment = CENT
        c = ws.cell(row=rr, column=4, value=f"='1. Inputs'!{col}{refs_inputs['ach']}"); c.font, c.number_format, c.border, c.alignment = BLK, PCT2, BORD, CENT
        c = ws.cell(row=rr, column=5, value=f"='1. Inputs'!{col}{refs_inputs['car']}"); c.font, c.number_format, c.border, c.alignment = BLK, PCT2, BORD, CENT
        c = ws.cell(row=rr, column=6, value=f'=IF(D{rr}="","",IF(ROUND(D{rr}-E{rr},6)=0,"NÃO","SIM"))'); c.font, c.border, c.alignment = BLK, BORD, CENT
        c = ws.cell(row=rr, column=7, value=f'=IFERROR(IF(D{rr}=0,"",E{rr}/D{rr}),"")'); c.font, c.number_format, c.border, c.alignment = BLK, PCT4, BORD, CENT
        c = ws.cell(row=rr, column=8, value=f'=IF(D{rr}="","(item vazio)",IF(F{rr}="NÃO","M1 — Gross-up sem redução","M2 — Gross-up com redução de base"))')
        c.font, c.border, c.alignment = BLK, BORD, CENT

    r = r + N_ITENS + 2

    ws.merge_cells(f"B{r}:H{r}")
    ws.cell(row=r, column=2, value="FÓRMULAS PRONTAS PARA O CONFIGURADOR DE TRIBUTOS").font = H2; r += 2

    # M1
    ws.merge_cells(f"B{r}:H{r}")
    c = ws.cell(row=r, column=2, value="MÉTODO 1 — Base do ICMS SEM redução de base (alíquota cheia)")
    c.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF"); c.fill = MID; r += 1
    ws.cell(row=r, column=2, value="Aplicação:").font = BOLD
    ws.merge_cells(f"C{r}:H{r}")
    ws.cell(row=r, column=3, value="itens com alíquota cheia (cartuchos, eletrônicos, mercadorias em geral).").font = BLK; r += 1
    ws.cell(row=r, column=2, value="Valor Origem:").font = BOLD
    ws.cell(row=r, column=3, value="11 — Fórmula Manual").font = BLK
    ws.merge_cells(f"C{r}:H{r}"); r += 1
    ws.cell(row=r, column=2, value="Adições/Deduções:").font = BOLD
    ws.cell(row=r, column=3, value='Frete = "3 - Soma"  |  Seguro = "3 - Soma"  |  Despesas = "3 - Soma"  |  Desconto = "2 - Subtrai"').font = BLK
    ws.merge_cells(f"C{r}:H{r}"); r += 1
    ws.cell(row=r, column=2, value="Fórmula:").font = BOLD
    ws.merge_cells(f"C{r}:H{r}")
    f1 = (f'="( O:VAL_MERCADORIA - O:DESCONTO + O:DESPESAS + O:FRETE + O:SEGURO + VAL:"'
          f'&{ref_cod["ii"]}&" + VAL:"&{ref_cod["ipi"]}'
          f'&" + VAL:"&{ref_cod["pis"]}&" + VAL:"&{ref_cod["cof"]}'
          f'&" ) / ( 1 - A:"&{ref_cod["icms"]}&" )"')
    c = ws.cell(row=r, column=3, value=f1); c.font, c.fill, c.border, c.alignment = MONO, CODEF, BORD, LEFT
    ws.row_dimensions[r].height = 28
    r += 2

    # M2
    ws.merge_cells(f"B{r}:H{r}")
    c = ws.cell(row=r, column=2, value="MÉTODO 2 — Base do ICMS COM redução de base (Conv. 52/91 e similares)")
    c.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF"); c.fill = MID; r += 1
    ws.cell(row=r, column=2, value="Aplicação:").font = BOLD
    ws.merge_cells(f"C{r}:H{r}")
    ws.cell(row=r, column=3, value="máquinas industriais e partes — NCMs 8443.39.10, 8443.91.99, 8428.33.00 etc.").font = BLK; r += 1
    ws.cell(row=r, column=2, value="Valor Origem:").font = BOLD
    ws.cell(row=r, column=3, value="11 — Fórmula Manual").font = BLK
    ws.merge_cells(f"C{r}:H{r}"); r += 1
    ws.cell(row=r, column=2, value="Adições/Deduções:").font = BOLD
    ws.cell(row=r, column=3, value='Frete = "3 - Soma"  |  Seguro = "3 - Soma"  |  Despesas = "3 - Soma"  |  Desconto = "2 - Subtrai"').font = BLK
    ws.merge_cells(f"C{r}:H{r}"); r += 1
    ws.cell(row=r, column=2, value="Fator de redução:").font = BOLD
    c = ws.cell(row=r, column=3, value=f"=IFERROR(ROUND('1. Inputs'!C{refs_inputs['car']}/'1. Inputs'!C{refs_inputs['ach']},4),0.4889)")
    c.font, c.number_format, c.border = BLK, "0.0000", BORD
    ws.cell(row=r, column=4, value="ex.: 8,8% ÷ 18% = 0,4889 (base reduzida para 48,89%)").font = Font(name=ARIAL, size=9, italic=True, color="595959")
    ws.merge_cells(f"D{r}:H{r}")
    r_fator = r  # guarda a referência da célula com o fator
    r += 1
    ws.cell(row=r, column=2, value="Fórmula:").font = BOLD
    ws.merge_cells(f"C{r}:H{r}")
    # Construímos o fator como string com vírgula sem depender de locale:
    # "INT(valor) & "," & TEXT(MOD(ROUND(valor*10000,0),10000),"0000")"
    # Para 0,4889 → "0" & "," & "4889" = "0,4889"
    # Para 1,0000 → "1" & "," & "0000" = "1,0000"
    f2 = (f'="( ( O:VAL_MERCADORIA - O:DESCONTO + O:DESPESAS + O:FRETE + O:SEGURO + VAL:"'
          f'&{ref_cod["ii"]}&" + VAL:"&{ref_cod["pis"]}&" + VAL:"&{ref_cod["cof"]}'
          f'&" ) / ( 1 - A:"&{ref_cod["icms"]}&" ) ) * "'
          f'&INT(C{r_fator})&","&TEXT(MOD(ROUND(C{r_fator}*10000,0),10000),"0000")')
    c = ws.cell(row=r, column=3, value=f2); c.font, c.fill, c.border, c.alignment = MONO, CODEF, BORD, LEFT
    ws.row_dimensions[r].height = 30
    r += 1

    ws.merge_cells(f"B{r}:H{r}")
    ws.cell(row=r, column=2, value="Obs.: o fator usa o Item 1 como base. Para reduções distintas no mesmo processo, replique a regra mudando apenas o multiplicador final. No M2, o IPI não entra no numerador (itens com redução geralmente têm IPI 0%); se o seu caso tiver IPI tributado com redução de base, acrescente \" + VAL:\"&{cod} antes do \") /\" para somar o IPI.").font = Font(name=ARIAL, size=9, italic=True, color="C00000")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 40
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 18


# ============================================================
# Aba 4: Conferência
# ============================================================
def _build_conferencia(wb: Workbook, refs_calc: dict, totais_espelho: dict) -> None:
    ws = wb.create_sheet("4. Conferência")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:G2")
    c = ws["B2"]; c.value = "CONFERÊNCIA — CÁLCULO × ESPELHO DA NF"
    c.font, c.fill, c.alignment = H1, NAVY, CENT
    ws.row_dimensions[2].height = 24
    ws.merge_cells("B3:G3")
    ws["B3"].value = "Os totais do espelho já foram preenchidos a partir do PDF. Ajuste se necessário; o status atualiza automaticamente."
    ws["B3"].font = Font(name=ARIAL, size=9, italic=True, color="595959")

    for j, h in enumerate(["Valor", "Calculado (planilha)", "Espelho da NF", "Diferença", "Status"], start=2):
        c = ws.cell(row=5, column=j, value=h); c.font, c.fill, c.alignment, c.border = HDR, MID, CENT, BORD

    linhas = [
        ("vProd total",                            f"='2. Cálculo'!{TOT_COL}{refs_calc['vp']}",   totais_espelho.get("vprod")),
        ("CIF / VMLD total",                       f"='2. Cálculo'!{TOT_COL}{refs_calc['cif']}",  totais_espelho.get("vcif")),
        ("II total",                               f"='2. Cálculo'!{TOT_COL}{refs_calc['ii']}",   totais_espelho.get("vii_total")),
        ("IPI total",                              f"='2. Cálculo'!{TOT_COL}{refs_calc['ipi']}",  totais_espelho.get("vipi")),
        ("PIS-importação total",                   f"='2. Cálculo'!{TOT_COL}{refs_calc['pis']}",  totais_espelho.get("vpis")),
        ("COFINS-importação total",                f"='2. Cálculo'!{TOT_COL}{refs_calc['cof']}",  totais_espelho.get("vcofins")),
        ("Outras desp. na base ICMS (Siscomex + AFRMM)",
                                                   f"='2. Cálculo'!{TOT_COL}{refs_calc['outd']}",
                                                   round(totais_espelho.get("sisc_total", 0) + totais_espelho.get("afrmm", 0), 2) or None),
        ("Base de cálculo do ICMS",                f"='2. Cálculo'!{TOT_COL}{refs_calc['bas']}",  totais_espelho.get("bc_icms")),
        ("ICMS destacado (a recolher)",            f"='2. Cálculo'!{TOT_COL}{refs_calc['icm']}",  totais_espelho.get("vicms")),
        ("ICMS 'valor devido' (Portal)",           f"='2. Cálculo'!{TOT_COL}{refs_calc['dev']}",  None),
        ("Outras desp. s/ influência no ICMS (Desp Ac − AFRMM)",
                                                   f"='2. Cálculo'!{TOT_COL}{refs_calc['outv']}",
            (round(totais_espelho.get("outras_despesas", 0) - totais_espelho.get("sisc_total", 0) - totais_espelho.get("afrmm", 0), 2)
             if totais_espelho.get("outras_despesas") else None)),
        ("vNF",                                    f"='2. Cálculo'!{TOT_COL}{refs_calc['vnf']}",  totais_espelho.get("vnf")),
    ]
    for i, (lab, calc, esp) in enumerate(linhas):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).font = BLK; ws.cell(row=r, column=2).border = BORD
        c = ws.cell(row=r, column=3, value=calc); c.font, c.number_format, c.border = BOLD, NUM, BORD
        c = ws.cell(row=r, column=4)
        if esp is not None:
            c.value = esp
        c.fill, c.font, c.number_format, c.border = INPF, BLU, NUM, BORD
        c = ws.cell(row=r, column=5, value=f'=IF(D{r}="","",D{r}-C{r})'); c.font, c.number_format, c.border = BLK, NUM, BORD
        c = ws.cell(row=r, column=6, value=f'=IF(D{r}="","(preencher)",IF(ROUND(E{r},2)=0,"OK ✓","DIVERGE"))')
        c.font, c.alignment, c.border = BLK, CENT, BORD

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14


# ============================================================
# Aba 5: Guia
# ============================================================
def _build_guia(wb: Workbook) -> None:
    ws = wb.create_sheet("5. Guia")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:G2")
    c = ws["B2"]; c.value = "GUIA RÁPIDO — IDENTIFICADORES E ARMADILHAS"
    c.font, c.fill, c.alignment = H1, NAVY, CENT
    ws.row_dimensions[2].height = 24

    def _sec(r, t):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        c = ws.cell(row=r, column=2, value=t)
        c.font, c.fill = Font(name=ARIAL, size=11, bold=True, color="FFFFFF"), MID
        ws.row_dimensions[r].height = 20

    def _tx(r, t, bold=False, mono=False):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        c = ws.cell(row=r, column=2, value=t)
        c.alignment = LEFT
        if bold: c.font = H2
        elif mono: c.font = MONO
        else: c.font = Font(name=ARIAL, size=10)
        ws.row_dimensions[r].height = 18

    r = 4
    _sec(r, "IDENTIFICADORES NAS FÓRMULAS DO CONFIGURADOR"); r += 1
    ops = [
        ("O:VAL_MERCADORIA", "Valor da Mercadoria do item."),
        ("O:DESCONTO", "Valor do Desconto do item."),
        ("O:FRETE", "Valor do Frete rateado no item."),
        ("O:SEGURO", "Valor do Seguro rateado no item."),
        ("O:DESPESAS", "Despesas Acessórias rateadas (Siscomex, AFRMM, demais)."),
        ("VAL:<cod>", "Valor calculado por outra Regra (ex.: VAL:EII001 = II)."),
        ("A:<cod>", "Alíquota cadastrada em outra Regra (ex.: A:AIC001 = ICMS)."),
    ]
    for op, desc in ops:
        ws.cell(row=r, column=2, value=op).font = MONO
        ws.cell(row=r, column=2).border = BORD
        c = ws.cell(row=r, column=3, value=desc); c.font, c.border, c.alignment = BLK, BORD, LEFT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        r += 1
    r += 1

    _sec(r, "ENCADEAMENTO DO CÁLCULO DA IMPORTAÇÃO"); r += 1
    for t in [
        "Passo 1: CIF/VMLD = FOB + Frete + Seguro (em R$).",
        "Passo 2: II = CIF × alíq. II  (0% se Ex-tarifário).",
        "Passo 3: Base IPI = CIF + II  →  IPI = Base IPI × alíq. IPI.",
        "Passo 4: PIS = CIF × 2,10%    |   COFINS = CIF × 9,65% (ou 10,25% majorada).",
        "Passo 5: Numerador ICMS = CIF + II + IPI + PIS + COFINS + Siscomex + AFRMM + outras integrantes.",
        "Passo 6: Base ICMS = Numerador ÷ (1 − alíq. interna cheia)   [gross-up 'por dentro'].",
        "Passo 7: ICMS = Base × carga efetiva. Se carga = cheia → sem redução; se carga < cheia → há redução.",
    ]:
        r = _tx(r, t) or r + 1
    r += 1

    _sec(r, "ARMADILHAS COMUNS EM IMPORTAÇÃO"); r += 1
    arm = [
        ("COFINS majorada", "Adicional de 1pp em alguns NCMs (§21 art. 8º, Lei 10.865/04 → 10,25%). Máquinas (8443.39.10) sim; cartuchos (8443.99.23) não."),
        ("AFRMM só marítimo", "Não há AFRMM em modal aéreo, rodoviário ou ferroviário."),
        ("Ex-tarifário", "Reduz II para 0% ou 2%. Cadastre no campo Ex-Tarifário do enquadramento por NCM."),
        ("Frete nacional × internacional", "O frete internacional já está no CIF/VMLD; o nacional é despesa separada e NÃO entra na base do ICMS-importação."),
        ("'Outras Despesas' do espelho", "Pode incluir itens que NÃO foram somados à base (armazenagem, frete nacional). Confira sempre a memória de cálculo."),
        ("DUIMP × NF — divergências", "Quando a base do ICMS no extrato diverge da NF, costuma ser parametrização do registro da DUIMP, não do Protheus."),
    ]
    for lab, desc in arm:
        c = ws.cell(row=r, column=2, value=lab); c.font, c.border, c.alignment = BOLD, BORD, LEFT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        c = ws.cell(row=r, column=4, value=desc); c.font, c.border, c.alignment = BLK, BORD, LEFT
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 36
        r += 1

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 24
    for col in "CDEFG": ws.column_dimensions[col].width = 16


# ============================================================
# Função principal
# ============================================================
def build_excel(dados: dict) -> bytes:
    """
    Gera o arquivo Excel a partir do dicionário do parser (parser.parse_pdf()).
    Retorna os bytes do XLSX, prontos para st.download_button.
    """
    wb = Workbook()
    refs_inputs = _build_inputs(wb, dados)
    refs_calc   = _build_calculo(wb, refs_inputs)
    _build_formulas(wb, refs_inputs)
    _build_conferencia(wb, refs_calc, dados.get("totais", {}))
    _build_guia(wb)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
